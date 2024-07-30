import openpyxl
import shutil
import openpyxl.styles
import re
from pprint import pprint
class Reporter():

    def __init__(self, xlsx_file):
        self.xlsx_file = xlsx_file
        shutil.copyfile("report_template.xlsx", xlsx_file)
        self.wb = openpyxl.load_workbook(xlsx_file)
        self.ws = self.wb.active
        self.font_family = 'Calibri'
        self.ws[f'B1'].font = openpyxl.styles.Font(name=self.font_family, sz=12, bold=True)
        self.ws[f'D1'].font = openpyxl.styles.Font(name=self.font_family, sz=12, bold=True)
        self.ws[f'G1'].font = openpyxl.styles.Font(name=self.font_family, sz=12, bold=True)
        
        
    def create_services_report(self, service_versions, active_service_vulnerabilities, possible_service_vulnerabilities):
        
        offset = 10
        NameLabel_cell = 3
        SeviceName_cell = 3
        VulnLabel_cell = 4
        CVE_cell = 5
        Exploitability_cell = 6
        Impact_cell = 7
        Version_cell = 8
        Severity_cell = 9
        StartingV_cell = 10
        EndingV_cell = 11

        for service_name in service_versions:
            if service_name in active_service_vulnerabilities and service_name in possible_service_vulnerabilities:
                self.ws[f'A{NameLabel_cell}'] = 'Service Name'
                self.ws[f'B{VulnLabel_cell}'] = 'Active Vulnerabilities'
                self.ws[f'C{VulnLabel_cell}'] = 'Possible Vulnerabilities'
                self.ws.merge_cells(f"B{NameLabel_cell}:C{NameLabel_cell}")
                self.ws[f'B{SeviceName_cell}'].value = service_name
                self.ws[f'A{SeviceName_cell}'].font = openpyxl.styles.Font(name=self.font_family, bold=True)
                self.ws[f'B{SeviceName_cell}'].alignment = openpyxl.styles.Alignment(horizontal='center')
                self.ws[f'A{NameLabel_cell}'].alignment = openpyxl.styles.Alignment(horizontal='center')
                self.ws[f'B{VulnLabel_cell}'].alignment = openpyxl.styles.Alignment(horizontal='center')
                self.ws[f'C{VulnLabel_cell}'].alignment = openpyxl.styles.Alignment(horizontal='center')

                temp_counter = 0
                for index in range(len(active_service_vulnerabilities[service_name])):

                    self.ws[f'B{CVE_cell}'] = active_service_vulnerabilities[service_name][index]["CVE"]
                    self.ws[f'B{Exploitability_cell}'] = active_service_vulnerabilities[service_name][index]["Exploitability Score"]
                    self.ws[f'B{Impact_cell}'] = active_service_vulnerabilities[service_name][index]["Impact Score"]
                    self.ws[f'B{Version_cell}'] = active_service_vulnerabilities[service_name][index]["Service Version"]
                    self.ws[f'B{Severity_cell}'] = active_service_vulnerabilities[service_name][index]["Severity"]
                    self.ws[f'B{StartingV_cell}'] = active_service_vulnerabilities[service_name][index]["Starting Version"]
                    self.ws[f'B{EndingV_cell}'] = active_service_vulnerabilities[service_name][index]["Ending Version"]

                    if active_service_vulnerabilities[service_name][index]["Severity"] == "HIGH":
                        self.ws[f'B{Severity_cell}'].fill = openpyxl.styles.PatternFill("solid", fgColor="d11818")
                    elif active_service_vulnerabilities[service_name][index]["Severity"] == "MEDIUM":
                        self.ws[f'B{Severity_cell}'].fill = openpyxl.styles.PatternFill("solid", fgColor="FFC300")
                    elif active_service_vulnerabilities[service_name][index]["Severity"] == "LOW":
                        self.ws[f'B{Severity_cell}'].fill = openpyxl.styles.PatternFill("solid", fgColor="ecec0a")

                    self.ws[f'B{SeviceName_cell}'].font = openpyxl.styles.Font(name=self.font_family, bold=True)
                    self.ws[f'B{VulnLabel_cell}'].font = openpyxl.styles.Font(name=self.font_family)
                    self.ws[f'B{CVE_cell}'].font = openpyxl.styles.Font(name=self.font_family)
                    self.ws[f'B{Exploitability_cell}'].font = openpyxl.styles.Font(name=self.font_family)
                    self.ws[f'B{Impact_cell}'].font = openpyxl.styles.Font(name=self.font_family)
                    self.ws[f'B{Version_cell}'].font = openpyxl.styles.Font(name=self.font_family)
                    self.ws[f'B{Severity_cell}'].font = openpyxl.styles.Font(name=self.font_family)
                    self.ws[f'B{StartingV_cell}'].font = openpyxl.styles.Font(name=self.font_family)
                    self.ws[f'B{EndingV_cell}'].font = openpyxl.styles.Font(name=self.font_family)
                    self.ws[f'A{CVE_cell}'] = "CVE"
                    self.ws[f'A{Exploitability_cell}'] = "Exploitability Score"
                    self.ws[f'A{Impact_cell}'] = "Impact Score"
                    self.ws[f'A{Version_cell}'] = "Service Version"
                    self.ws[f'A{Severity_cell}'] = "Severity"
                    self.ws[f'A{StartingV_cell}'] = "Starting version"
                    self.ws[f'A{EndingV_cell}'] = "Ending Version"
                    self.ws[f'A{CVE_cell}'].font = openpyxl.styles.Font(name=self.font_family, bold=True)
                    self.ws[f'A{Exploitability_cell}'].font = openpyxl.styles.Font(name=self.font_family, bold=True)
                    self.ws[f'A{Impact_cell}'].font = openpyxl.styles.Font(name=self.font_family, bold=True)
                    self.ws[f'A{Version_cell}'].font = openpyxl.styles.Font(name=self.font_family, bold=True)
                    self.ws[f'A{Severity_cell}'].font = openpyxl.styles.Font(name=self.font_family, bold=True)
                    self.ws[f'A{StartingV_cell}'].font = openpyxl.styles.Font(name=self.font_family, bold=True)
                    self.ws[f'A{EndingV_cell}'].font = openpyxl.styles.Font(name=self.font_family, bold=True)
                    self.ws[f'A{SeviceName_cell}'].alignment = openpyxl.styles.Alignment(horizontal='right')
                    self.ws[f'A{CVE_cell}'].alignment = openpyxl.styles.Alignment(horizontal='right')
                    self.ws[f'A{Exploitability_cell}'].alignment = openpyxl.styles.Alignment(horizontal='right')
                    self.ws[f'A{Impact_cell}'].alignment = openpyxl.styles.Alignment(horizontal='right')
                    self.ws[f'A{Version_cell}'].alignment = openpyxl.styles.Alignment(horizontal='right')
                    self.ws[f'A{Severity_cell}'].alignment = openpyxl.styles.Alignment(horizontal='right')
                    self.ws[f'A{StartingV_cell}'].alignment = openpyxl.styles.Alignment(horizontal='right')
                    self.ws[f'A{EndingV_cell}'].alignment = openpyxl.styles.Alignment(horizontal='right')
                    self.ws[f'B{CVE_cell}'].alignment = openpyxl.styles.Alignment(horizontal='left')
                    self.ws[f'B{Exploitability_cell}'].alignment = openpyxl.styles.Alignment(horizontal='left')
                    self.ws[f'B{Impact_cell}'].alignment = openpyxl.styles.Alignment(horizontal='left')
                    self.ws[f'B{Version_cell}'].alignment = openpyxl.styles.Alignment(horizontal='left')
                    self.ws[f'B{Severity_cell}'].alignment = openpyxl.styles.Alignment(horizontal='left')

                    NameLabel_cell += offset
                    VulnLabel_cell += offset
                    SeviceName_cell += offset
                    CVE_cell += offset
                    Exploitability_cell += offset
                    Impact_cell += offset
                    Version_cell += offset
                    Severity_cell += offset
                    StartingV_cell += offset
                    EndingV_cell += offset
                    temp_counter+=1
                temp_offset = temp_counter*offset
                
                for index in range(len(possible_service_vulnerabilities[service_name])):
                    
                    self.ws[f'C{CVE_cell-temp_offset}'] = possible_service_vulnerabilities[service_name][index]["CVE"]
                    self.ws[f'C{Exploitability_cell-temp_offset}'] = possible_service_vulnerabilities[service_name][index]["Exploitability Score"]
                    self.ws[f'C{Impact_cell-temp_offset}'] = possible_service_vulnerabilities[service_name][index]["Impact Score"]
                    self.ws[f'C{Version_cell-temp_offset}'] = possible_service_vulnerabilities[service_name][index]["Service Version"]
                    self.ws[f'C{Severity_cell-temp_offset}'] = possible_service_vulnerabilities[service_name][index]["Severity"]
                    self.ws[f'C{StartingV_cell-temp_offset}'] = possible_service_vulnerabilities[service_name][index]["Starting Version"]
                    self.ws[f'C{EndingV_cell-temp_offset}'] = possible_service_vulnerabilities[service_name][index]["Ending Version"]

                    if possible_service_vulnerabilities[service_name][index]["Severity"] == "HIGH":
                        self.ws[f'C{Severity_cell-temp_offset}'].fill = openpyxl.styles.PatternFill("solid", fgColor="d11818")
                    elif possible_service_vulnerabilities[service_name][index]["Severity"] == "MEDIUM":
                        self.ws[f'C{Severity_cell-temp_offset}'].fill = openpyxl.styles.PatternFill("solid", fgColor="FFC300")
                    elif possible_service_vulnerabilities[service_name][index]["Severity"] == "LOW":
                        self.ws[f'C{Severity_cell-temp_offset}'].fill = openpyxl.styles.PatternFill("solid", fgColor="ecec0a")

                    self.ws[f'C{SeviceName_cell-temp_offset}'].font = openpyxl.styles.Font(name=self.font_family, bold=True)
                    self.ws[f'C{VulnLabel_cell-temp_offset}'].font = openpyxl.styles.Font(name=self.font_family)
                    self.ws[f'C{CVE_cell-temp_offset}'].font = openpyxl.styles.Font(name=self.font_family)
                    self.ws[f'C{Exploitability_cell-temp_offset}'].font = openpyxl.styles.Font(name=self.font_family)
                    self.ws[f'C{Impact_cell-temp_offset}'].font = openpyxl.styles.Font(name=self.font_family)
                    self.ws[f'C{Version_cell-temp_offset}'].font = openpyxl.styles.Font(name=self.font_family)
                    self.ws[f'C{Severity_cell-temp_offset}'].font = openpyxl.styles.Font(name=self.font_family)
                    self.ws[f'C{StartingV_cell-temp_offset}'].font = openpyxl.styles.Font(name=self.font_family)
                    self.ws[f'C{EndingV_cell-temp_offset}'].font = openpyxl.styles.Font(name=self.font_family)
                    self.ws[f'A{CVE_cell-temp_offset}'] = "CVE"
                    self.ws[f'A{Exploitability_cell-temp_offset}'] = "Exploitability Score"
                    self.ws[f'A{Impact_cell-temp_offset}'] = "Impact Score"
                    self.ws[f'A{Version_cell-temp_offset}'] = "Service Version"
                    self.ws[f'A{Severity_cell-temp_offset}'] = "Severity"
                    self.ws[f'A{StartingV_cell-temp_offset}'] = "Starting version"
                    self.ws[f'A{EndingV_cell-temp_offset}'] = "Ending Version"
                    self.ws[f'A{CVE_cell-temp_offset}'].font = openpyxl.styles.Font(name=self.font_family, bold=True)
                    self.ws[f'A{Exploitability_cell-temp_offset}'].font = openpyxl.styles.Font(name=self.font_family, bold=True)
                    self.ws[f'A{Impact_cell-temp_offset}'].font = openpyxl.styles.Font(name=self.font_family, bold=True)
                    self.ws[f'A{Version_cell-temp_offset}'].font = openpyxl.styles.Font(name=self.font_family, bold=True)
                    self.ws[f'A{Severity_cell-temp_offset}'].font = openpyxl.styles.Font(name=self.font_family, bold=True)
                    self.ws[f'A{StartingV_cell-temp_offset}'].font = openpyxl.styles.Font(name=self.font_family, bold=True)
                    self.ws[f'A{EndingV_cell-temp_offset}'].font = openpyxl.styles.Font(name=self.font_family, bold=True)
                    self.ws[f'A{SeviceName_cell-temp_offset}'].alignment = openpyxl.styles.Alignment(horizontal='right')
                    self.ws[f'A{CVE_cell-temp_offset}'].alignment = openpyxl.styles.Alignment(horizontal='right')
                    self.ws[f'A{Exploitability_cell-temp_offset}'].alignment = openpyxl.styles.Alignment(horizontal='right')
                    self.ws[f'A{Impact_cell-temp_offset}'].alignment = openpyxl.styles.Alignment(horizontal='right')
                    self.ws[f'A{Version_cell-temp_offset}'].alignment = openpyxl.styles.Alignment(horizontal='right')
                    self.ws[f'A{Severity_cell-temp_offset}'].alignment = openpyxl.styles.Alignment(horizontal='right')
                    self.ws[f'A{StartingV_cell-temp_offset}'].alignment = openpyxl.styles.Alignment(horizontal='right')
                    self.ws[f'A{EndingV_cell-temp_offset}'].alignment = openpyxl.styles.Alignment(horizontal='right')
                    self.ws[f'C{CVE_cell-temp_offset}'].alignment = openpyxl.styles.Alignment(horizontal='left')
                    self.ws[f'C{Exploitability_cell-temp_offset}'].alignment = openpyxl.styles.Alignment(horizontal='left')
                    self.ws[f'C{Impact_cell-temp_offset}'].alignment = openpyxl.styles.Alignment(horizontal='left')
                    self.ws[f'C{Version_cell-temp_offset}'].alignment = openpyxl.styles.Alignment(horizontal='left')
                    self.ws[f'C{Severity_cell-temp_offset}'].alignment = openpyxl.styles.Alignment(horizontal='left')

                    temp_offset -=offset

                    
            elif service_name in active_service_vulnerabilities:
                self.ws[f'A{NameLabel_cell}'] = 'Service Name'
                self.ws[f'B{VulnLabel_cell}'] = 'Active Vulnerabilities'
                self.ws[f'C{VulnLabel_cell}'] = 'Possible Vulnerabilities'
                self.ws.merge_cells(f"B{NameLabel_cell}:C{NameLabel_cell}")
                self.ws[f'B{SeviceName_cell}'].value = service_name
                self.ws[f'A{SeviceName_cell}'].font = openpyxl.styles.Font(name=self.font_family, bold=True)
                self.ws[f'B{SeviceName_cell}'].alignment = openpyxl.styles.Alignment(horizontal='center')
                self.ws[f'A{NameLabel_cell}'].alignment = openpyxl.styles.Alignment(horizontal='center')
                self.ws[f'B{VulnLabel_cell}'].alignment = openpyxl.styles.Alignment(horizontal='center')
                self.ws[f'C{VulnLabel_cell}'].alignment = openpyxl.styles.Alignment(horizontal='center')

                for index in range(len(active_service_vulnerabilities[service_name])):

                    self.ws[f'B{CVE_cell}'] = active_service_vulnerabilities[service_name][index]["CVE"]
                    self.ws[f'B{Exploitability_cell}'] = active_service_vulnerabilities[service_name][index]["Exploitability Score"]
                    self.ws[f'B{Impact_cell}'] = active_service_vulnerabilities[service_name][index]["Impact Score"]
                    self.ws[f'B{Version_cell}'] = active_service_vulnerabilities[service_name][index]["Service Version"]
                    self.ws[f'B{Severity_cell}'] = active_service_vulnerabilities[service_name][index]["Severity"]
                    self.ws[f'B{StartingV_cell}'] = active_service_vulnerabilities[service_name][index]["Starting Version"]
                    self.ws[f'B{EndingV_cell}'] = active_service_vulnerabilities[service_name][index]["Ending Version"]

                    if possible_service_vulnerabilities[service_name][index]["Severity"] == "CRITICAL":
                        self.ws[f'C{Severity_cell}'].fill = openpyxl.styles.PatternFill("solid", fgColor="830462")
                    elif possible_service_vulnerabilities[service_name][index]["Severity"] == "HIGH":
                        self.ws[f'B{Severity_cell}'].fill = openpyxl.styles.PatternFill("solid", fgColor="d11818")
                    elif active_service_vulnerabilities[service_name][index]["Severity"] == "MEDIUM":
                        self.ws[f'B{Severity_cell}'].fill = openpyxl.styles.PatternFill("solid", fgColor="FFC300")
                    elif active_service_vulnerabilities[service_name][index]["Severity"] == "LOW":
                        self.ws[f'B{Severity_cell}'].fill = openpyxl.styles.PatternFill("solid", fgColor="ecec0a")

                    self.ws[f'B{SeviceName_cell}'].font = openpyxl.styles.Font(name=self.font_family, bold=True)
                    self.ws[f'B{VulnLabel_cell}'].font = openpyxl.styles.Font(name=self.font_family)
                    self.ws[f'B{CVE_cell}'].font = openpyxl.styles.Font(name=self.font_family)
                    self.ws[f'B{Exploitability_cell}'].font = openpyxl.styles.Font(name=self.font_family)
                    self.ws[f'B{Impact_cell}'].font = openpyxl.styles.Font(name=self.font_family)
                    self.ws[f'B{Version_cell}'].font = openpyxl.styles.Font(name=self.font_family)
                    self.ws[f'B{Severity_cell}'].font = openpyxl.styles.Font(name=self.font_family)
                    self.ws[f'B{StartingV_cell}'].font = openpyxl.styles.Font(name=self.font_family)
                    self.ws[f'B{EndingV_cell}'].font = openpyxl.styles.Font(name=self.font_family)
                    self.ws[f'A{CVE_cell}'] = "CVE"
                    self.ws[f'A{Exploitability_cell}'] = "Exploitability Score"
                    self.ws[f'A{Impact_cell}'] = "Impact Score"
                    self.ws[f'A{Version_cell}'] = "Service Version"
                    self.ws[f'A{Severity_cell}'] = "Severity"
                    self.ws[f'A{StartingV_cell}'] = "Starting version"
                    self.ws[f'A{EndingV_cell}'] = "Ending Version"
                    self.ws[f'A{CVE_cell}'].font = openpyxl.styles.Font(name=self.font_family, bold=True)
                    self.ws[f'A{Exploitability_cell}'].font = openpyxl.styles.Font(name=self.font_family, bold=True)
                    self.ws[f'A{Impact_cell}'].font = openpyxl.styles.Font(name=self.font_family, bold=True)
                    self.ws[f'A{Version_cell}'].font = openpyxl.styles.Font(name=self.font_family, bold=True)
                    self.ws[f'A{Severity_cell}'].font = openpyxl.styles.Font(name=self.font_family, bold=True)
                    self.ws[f'A{StartingV_cell}'].font = openpyxl.styles.Font(name=self.font_family, bold=True)
                    self.ws[f'A{EndingV_cell}'].font = openpyxl.styles.Font(name=self.font_family, bold=True)
                    self.ws[f'A{SeviceName_cell}'].alignment = openpyxl.styles.Alignment(horizontal='right')
                    self.ws[f'A{CVE_cell}'].alignment = openpyxl.styles.Alignment(horizontal='right')
                    self.ws[f'A{Exploitability_cell}'].alignment = openpyxl.styles.Alignment(horizontal='right')
                    self.ws[f'A{Impact_cell}'].alignment = openpyxl.styles.Alignment(horizontal='right')
                    self.ws[f'A{Version_cell}'].alignment = openpyxl.styles.Alignment(horizontal='right')
                    self.ws[f'A{Severity_cell}'].alignment = openpyxl.styles.Alignment(horizontal='right')
                    self.ws[f'A{StartingV_cell}'].alignment = openpyxl.styles.Alignment(horizontal='right')
                    self.ws[f'A{EndingV_cell}'].alignment = openpyxl.styles.Alignment(horizontal='right')
                    self.ws[f'B{CVE_cell}'].alignment = openpyxl.styles.Alignment(horizontal='left')
                    self.ws[f'B{Exploitability_cell}'].alignment = openpyxl.styles.Alignment(horizontal='left')
                    self.ws[f'B{Impact_cell}'].alignment = openpyxl.styles.Alignment(horizontal='left')
                    self.ws[f'B{Version_cell}'].alignment = openpyxl.styles.Alignment(horizontal='left')
                    self.ws[f'B{Severity_cell}'].alignment = openpyxl.styles.Alignment(horizontal='left')

                    NameLabel_cell += offset
                    VulnLabel_cell += offset
                    SeviceName_cell += offset
                    CVE_cell += offset
                    Exploitability_cell += offset
                    Impact_cell += offset
                    Version_cell += offset
                    Severity_cell += offset
                    StartingV_cell += offset
                    EndingV_cell += offset

            elif service_name in possible_service_vulnerabilities:
                self.ws[f'A{NameLabel_cell}'] = 'Service Name'
                self.ws[f'B{VulnLabel_cell}'] = 'Active Vulnerabilities'
                self.ws[f'C{VulnLabel_cell}'] = 'Possible Vulnerabilities'
                self.ws.merge_cells(f"B{NameLabel_cell}:C{NameLabel_cell}")
                self.ws[f'B{SeviceName_cell}'].value = service_name
                self.ws[f'A{SeviceName_cell}'].font = openpyxl.styles.Font(name=self.font_family, bold=True)
                self.ws[f'C{VulnLabel_cell}'].alignment = openpyxl.styles.Alignment(horizontal='center')
                self.ws[f'B{SeviceName_cell}'].alignment = openpyxl.styles.Alignment(horizontal='center')
                self.ws[f'A{NameLabel_cell}'].alignment = openpyxl.styles.Alignment(horizontal='center')
                self.ws[f'B{VulnLabel_cell}'].alignment = openpyxl.styles.Alignment(horizontal='center')
                
                for index in range(len(possible_service_vulnerabilities[service_name])):
                    self.ws[f'C{CVE_cell}'] = possible_service_vulnerabilities[service_name][index]["CVE"]
                    self.ws[f'C{Exploitability_cell}'] = possible_service_vulnerabilities[service_name][index]["Exploitability Score"]
                    self.ws[f'C{Impact_cell}'] = possible_service_vulnerabilities[service_name][index]["Impact Score"]
                    self.ws[f'C{Version_cell}'] = possible_service_vulnerabilities[service_name][index]["Service Version"]
                    self.ws[f'C{Severity_cell}'] = possible_service_vulnerabilities[service_name][index]["Severity"]
                    self.ws[f'C{StartingV_cell}'] = possible_service_vulnerabilities[service_name][index]["Starting Version"]
                    self.ws[f'C{EndingV_cell}'] = possible_service_vulnerabilities[service_name][index]["Ending Version"]
                    self.ws[f'A{SeviceName_cell}'].alignment = openpyxl.styles.Alignment(horizontal='right')


                    if possible_service_vulnerabilities[service_name][index]["Severity"] == "CRITICAL":
                        self.ws[f'C{Severity_cell}'].fill = openpyxl.styles.PatternFill("solid", fgColor="830462")
                    elif possible_service_vulnerabilities[service_name][index]["Severity"] == "HIGH":
                        self.ws[f'C{Severity_cell}'].fill = openpyxl.styles.PatternFill("solid", fgColor="d11818")
                    elif possible_service_vulnerabilities[service_name][index]["Severity"] == "MEDIUM":
                        self.ws[f'C{Severity_cell}'].fill = openpyxl.styles.PatternFill("solid", fgColor="FFC300")
                    elif possible_service_vulnerabilities[service_name][index]["Severity"] == "LOW":
                        self.ws[f'C{Severity_cell}'].fill = openpyxl.styles.PatternFill("solid", fgColor="ecec0a")
                    
                    self.ws[f'A{CVE_cell}'] = "CVE"
                    self.ws[f'A{Exploitability_cell}'] = "Exploitability Score"
                    self.ws[f'A{Impact_cell}'] = "Impact Score"
                    self.ws[f'A{Version_cell}'] = "Service Version"
                    self.ws[f'A{Severity_cell}'] = "Severity"
                    self.ws[f'A{StartingV_cell}'] = "Starting version"
                    self.ws[f'A{EndingV_cell}'] = "Ending Version"
                    self.ws[f'A{CVE_cell}'].font = openpyxl.styles.Font(name=self.font_family, bold=True)
                    self.ws[f'A{Exploitability_cell}'].font = openpyxl.styles.Font(name=self.font_family, bold=True)
                    self.ws[f'A{Impact_cell}'].font = openpyxl.styles.Font(name=self.font_family, bold=True)
                    self.ws[f'A{Version_cell}'].font = openpyxl.styles.Font(name=self.font_family, bold=True)
                    self.ws[f'A{Severity_cell}'].font = openpyxl.styles.Font(name=self.font_family, bold=True)
                    self.ws[f'A{StartingV_cell}'].font = openpyxl.styles.Font(name=self.font_family, bold=True)
                    self.ws[f'A{EndingV_cell}'].font = openpyxl.styles.Font(name=self.font_family, bold=True)
                    self.ws[f'A{SeviceName_cell}'].alignment = openpyxl.styles.Alignment(horizontal='right')
                    self.ws[f'A{CVE_cell}'].alignment = openpyxl.styles.Alignment(horizontal='right')
                    self.ws[f'A{Exploitability_cell}'].alignment = openpyxl.styles.Alignment(horizontal='right')
                    self.ws[f'A{Impact_cell}'].alignment = openpyxl.styles.Alignment(horizontal='right')
                    self.ws[f'A{Version_cell}'].alignment = openpyxl.styles.Alignment(horizontal='right')
                    self.ws[f'A{Severity_cell}'].alignment = openpyxl.styles.Alignment(horizontal='right')
                    self.ws[f'A{StartingV_cell}'].alignment = openpyxl.styles.Alignment(horizontal='right')
                    self.ws[f'A{EndingV_cell}'].alignment = openpyxl.styles.Alignment(horizontal='right')
                    self.ws[f'C{VulnLabel_cell}'].font = openpyxl.styles.Font(name=self.font_family)
                    self.ws[f'C{CVE_cell}'].font = openpyxl.styles.Font(name=self.font_family)
                    self.ws[f'C{Exploitability_cell}'].font = openpyxl.styles.Font(name=self.font_family)
                    self.ws[f'C{Impact_cell}'].font = openpyxl.styles.Font(name=self.font_family)
                    self.ws[f'C{Version_cell}'].font = openpyxl.styles.Font(name=self.font_family)
                    self.ws[f'C{Severity_cell}'].font = openpyxl.styles.Font(name=self.font_family)
                    self.ws[f'C{StartingV_cell}'].font = openpyxl.styles.Font(name=self.font_family)
                    self.ws[f'C{EndingV_cell}'].font = openpyxl.styles.Font(name=self.font_family)
                    self.ws[f'C{CVE_cell}'].alignment = openpyxl.styles.Alignment(horizontal='left')
                    self.ws[f'C{Exploitability_cell}'].alignment = openpyxl.styles.Alignment(horizontal='left')
                    self.ws[f'C{Impact_cell}'].alignment = openpyxl.styles.Alignment(horizontal='left')
                    self.ws[f'C{Version_cell}'].alignment = openpyxl.styles.Alignment(horizontal='left')
                    self.ws[f'C{Severity_cell}'].alignment = openpyxl.styles.Alignment(horizontal='left')
                
                    NameLabel_cell += offset
                    VulnLabel_cell += offset
                    SeviceName_cell += offset
                    CVE_cell += offset
                    Exploitability_cell += offset
                    Impact_cell += offset
                    Version_cell += offset
                    Severity_cell += offset
                    StartingV_cell += offset
                    EndingV_cell += offset

        # Save the workbook
        self.wb.save(self.xlsx_file)
    
    def create_user_report(self, vulnerable_users, high_privilaged_users):
        starting_cell = 4
        self.ws[f'F{starting_cell-1}'].font = openpyxl.styles.Font(name=self.font_family, bold=True)
        self.ws[f'E{starting_cell-1}'].font = openpyxl.styles.Font(name=self.font_family, bold=True)

        for user in high_privilaged_users:
            self.ws[f'F{starting_cell}'] = f"{user} member of {high_privilaged_users[user]}"
            self.ws[f'F{starting_cell}'].font = openpyxl.styles.Font(name=self.font_family)
            self.ws[f'F{starting_cell}'].fill = openpyxl.styles.PatternFill("solid", fgColor="d11818")
            starting_cell += 1
        
        starting_cell = 4
        for user in vulnerable_users:
            if not user in high_privilaged_users:
                self.ws[f'E{starting_cell}'] = user
                self.ws[f'E{starting_cell}'].font = openpyxl.styles.Font(name=self.font_family)
                self.ws[f'E{starting_cell}'].fill = openpyxl.styles.PatternFill("solid", fgColor="ecec0a")
                starting_cell += 1
        
        self.wb.save(self.xlsx_file)

    def create_conf_report(self, configurations):
        recom_cell = 6
        serviceName_cell = recom_cell-3
        confFile_cell = recom_cell-2
        
        self.ws[f'H{serviceName_cell}'].font = openpyxl.styles.Font(name=self.font_family, bold=True)
        self.ws[f'I{serviceName_cell}'].font = openpyxl.styles.Font(name=self.font_family, bold=True)
        self.ws[f'H{confFile_cell}'].font = openpyxl.styles.Font(name=self.font_family, bold=True)
        self.ws[f'I{confFile_cell}'].font = openpyxl.styles.Font(name=self.font_family, bold=True)
        self.ws[f'I{recom_cell-1}'].font = openpyxl.styles.Font(name=self.font_family, bold=True)
        for service in configurations:
            if configurations[service]:
                serviceName_cell = recom_cell-3
                confFile_cell = recom_cell-2
                self.ws[f'I{confFile_cell-1}'] = service
                self.ws[f'H{confFile_cell-1}'] = "Service"
                self.ws[f'H{confFile_cell-1}'].alignment = openpyxl.styles.Alignment(horizontal='right')
                self.ws[f'I{confFile_cell-1}'].font = openpyxl.styles.Font(name=self.font_family, bold=True)
                self.ws[f'H{confFile_cell-1}'].font = openpyxl.styles.Font(name=self.font_family, bold=True)
                if service in ["Apache", "PostgreSQL", "Filezilla"]:
                    for file in configurations[service]:
                        if configurations[service][file]:
                            self.ws[f'H{confFile_cell}'] = "Configuration File"
                            self.ws[f'I{confFile_cell}'] = file
                            self.ws[f'I{recom_cell-1}'] = "Recommendations"
                            self.ws[f'H{confFile_cell}'].font = openpyxl.styles.Font(name=self.font_family, bold=True)
                            self.ws[f'I{confFile_cell}'].font = openpyxl.styles.Font(name=self.font_family, bold=True)
                            self.ws[f'I{recom_cell-1}'].font = openpyxl.styles.Font(name=self.font_family, bold=True)

                            for configuration in configurations[service][file].values():
                                if not isinstance(configuration, bool):
                                    if "Warning" in configuration:
                                        self.ws[f'I{recom_cell}'].fill = openpyxl.styles.PatternFill("solid", fgColor="ecec0a")

                                    self.ws[f'I{recom_cell}'] = configuration
                                    self.ws[f'I{recom_cell}'].font = openpyxl.styles.Font(name=self.font_family)
                                    recom_cell += 1
                            recom_cell += 2
                            confFile_cell = recom_cell-2
                elif service == "Registry":
                    for reg_key in configurations[service]:
                        if configurations[service][reg_key]:
                            index = 0
                            self.ws[f'H{recom_cell-2}'] = "Registry Key"
                            self.ws[f'I{recom_cell-2}'] = reg_key
                            self.ws[f'I{recom_cell-1}'] = "Needs review"
                            self.ws[f'H{recom_cell-2}'].font = openpyxl.styles.Font(name=self.font_family, bold=True)
                            self.ws[f'I{recom_cell-2}'].font = openpyxl.styles.Font(name=self.font_family, bold=True)
                            self.ws[f'I{recom_cell-1}'].font = openpyxl.styles.Font(name=self.font_family, bold=True)

                            while index < len(configurations[service][reg_key]):
                                if configurations[service][reg_key][index]:
                                    name = re.split('#', configurations[service][reg_key][index])
                                    if name[0] == "EnableFirewall" and name[1] == "0":
                                        self.ws[f'H{recom_cell}'].fill = openpyxl.styles.PatternFill("solid", fgColor="ecec0a")
                                        self.ws[f'I{recom_cell}'].fill = openpyxl.styles.PatternFill("solid", fgColor="ecec0a")
                                    self.ws[f'H{recom_cell}'] = name[0]
                                    self.ws[f'I{recom_cell}'] = name[1]
                                    self.ws[f'H{recom_cell}'].font = openpyxl.styles.Font(name=self.font_family)
                                    self.ws[f'I{recom_cell}'].font = openpyxl.styles.Font(name=self.font_family)
                                    recom_cell += 1
                                index += 1
                        recom_cell += 3
                        confFile_cell = recom_cell-2
                            
                elif service == "Nftables":
                    self.ws[f'I{recom_cell-3}'] = "Nftables"
                    self.ws[f'I{recom_cell-2}'] = "Needs review"
                    self.ws[f'H{recom_cell-2}'] = ""
                    self.ws[f'I{recom_cell-2}'].font = openpyxl.styles.Font(name=self.font_family, bold=True)
                    self.ws[f'I{recom_cell-3}'].font = openpyxl.styles.Font(name=self.font_family, bold=True)
                    self.ws[f'I{recom_cell-2}'].alignment = openpyxl.styles.Alignment(horizontal='center')

                    if configurations[service][0]:
                        if configurations[service][1]:
                            for rule in configurations[service][1]:
                                self.ws[f'I{recom_cell-1}'] = rule
                                recom_cell += 1
                    else:
                        self.ws[f'I{recom_cell-3}'] = "Nftables are inactive"
                        self.ws[f'I{recom_cell-3}'].font = openpyxl.styles.Font(name=self.font_family, bold=True)
                        self.ws[f'I{recom_cell-3}'].fill = openpyxl.styles.PatternFill("solid", fgColor="d11818")
                        self.ws[f'I{recom_cell}'] = "Start the nftables service"
                recom_cell += 2
        self.wb.save(self.xlsx_file)

    def xlsx_to_pdf(self, pdf_file, os):
        if os == "windows":
            import aspose.cells as ac
            workbook = ac.Workbook(self.xlsx_file)
            pdfOptions = ac.PdfSaveOptions()
            pdfOptions.all_columns_in_one_page_per_sheet = True
            workbook.save(pdf_file, pdfOptions)
            print(f"Report files {self.xlsx_file} and {pdf_file} have been created")
        else:
            print(f"Report files {self.xlsx_file} have been created")